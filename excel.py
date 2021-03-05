import openpyxl

workbook = openpyxl.load_workbook('torihiki.xlsx')
sheet = workbook["Sheet1"]

suppliers = []
transaction_amounts = []


for i in range(3,9):
    cell_value = sheet.cell(row=i,column=2).value

    if cell_value not in suppliers:
        suppliers.append(cell_value)

print(suppliers)

for i in suppliers:
    transaction_amount = 0
    for j in range(3,9):
        suppliers = sheet.cell(row=j,column=2).value

        if suppliers == i :
            transaction_amount += sheet.cell(row=j,column=6).value

    transaction_amounts.append(transaction_amount)

print(transaction_amounts)